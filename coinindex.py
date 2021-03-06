""" Choose your portfolio of cryptocurrencies """
""" Program calculates how many units you should buy """
""" to have coins proportional to their market cap """
""" which is similar to Index investing """

import os, json, urllib.request
from openpyxl import Workbook, load_workbook




def display_menu():
	print("\n\n\n   1. Choose your portfolio")
	print("   2. Update your spreadsheet")
	print("   3. Exit")
def display_top():
	while True:
		x=input("Display 30 cryptocurrencies beginning with rank... : ")
		try:
			int(x)
		except:
			print("Wrong input! Try again")
			continue
		if int(x)<len(data)-30 and x.isdigit() and int(x)>0:
			break
		else:
			print("Wrong input! Try again")

	for i in range(int(x)-1,int(x)+29):
		try:
			print ("{:>5}. {:>23}  {:>12} USD  Market Cap: {:>15,} USD".format(data[i]['rank'],data[i]['name'],data[i]['price_usd'],float(data[i]['market_cap_usd'])))
		except:
			print ("Can't display those cryptocurrencies. Market cap unknown...")
			return
	
def choose_portfolio():
	
	new_list=[]
	n_list=[]
	list_=[]
	rng=[]
	error=False

	display_top()
	

	while True:
		choice=input("\n  Choose your coins, by typing numbers separated by commas and hyphens i.e.: 1,3-5,15 : ")
		choice_list=choice.split(',')

		#checking if there's range in input and convert it (3-7) --> 3,4,5,6,7

		for item in choice_list:
			for i in range(len(item)):
				if item[i]=='-':
					rng=item.split('-')
					list_.extend([x for x in range(int(rng[0]),int(rng[1])+1)])
					
		#n_list filters choice_list removing items with hyphens and other non digit 

		
		for item in choice_list:
			if item.isdigit():
				n_list.append(item)

		# changes applied fo choice_list  1,3-5,7,11-15 ---> 1,3,4,5,7,11,12,13,14,15

		choice_list=[]
		choice_list.extend(list_)
		choice_list.extend(n_list)

		
		for i in choice_list:
			try:
				check=data[int(i)-1]      #checking if input is correct
			except:
				error=True
		if error:
			error=False
			print("\nInput error! Try again")
		else:
			break
		
	choice_list=set(choice_list) #removing duplicates
	choice_list=list(choice_list) #converting back to list so it can be sorted
	choice_list=sorted([int(i) for i in choice_list])

	
	for item in choice_list:
		new_list.append(data[item-1])
	
	print ("\nYour portfolio:\n")
	
	for item in new_list:
		print ("{:>5}. {:>23}  {:>12} USD  Market Cap: {:>15,} USD".format(item['rank'],
		item['name'],item['price_usd'],float(item['market_cap_usd'])))

	money=float(input("\nHow much money would you like to invest (in USD) ? "))
	total_market_cap=0

	for item in new_list:
		total_market_cap+=float(item['market_cap_usd'])
	
	for item in new_list:
		item['proportion']=float(item['market_cap_usd'])/total_market_cap
		item['worth_in_usd']=item['proportion']*money
		item['units']=item['worth_in_usd']/float(item['price_usd'])
		item['in_btc']=item['worth_in_usd']/float(data[0]['price_usd']) #bitcoin price in USD
		
	print ("\nTo have your investment proportional to crypto market cap your portfolio should look like this: \n")
	
	for item in new_list:
		print ("{:>5}. {:>23}  {:>10.3f} units {:>12.2f} in USD {:>10.4f} in BTC".format(item['rank'],
		item['name'],item['units'],item['worth_in_usd'],item['in_btc']))

	print("\n  1. Save portfolio to text file")
	print("  2. Create .xls file")
	print("  3. Exit")
	
	while True:
		option=input("\nChoose option: ")
	
		if option=='1' : 
			save_to_text_file(new_list)
			break
		elif option=='2' : 
			create_xls(new_list,money)
			break
		elif option=='3' : break
		else:
			print ("Invalid option!")
			
	
def save_to_text_file(data):
	file_name=input('Save portfolio as...: ')
	with open(file_name,'w') as f:
		for item in data:
			f.write("{:>4s}. {:>20s}  {:>15.3f} units {:>15.2f} in USD {:>15.4f} in BTC\n".format(item['rank'],
			item['name'],item['units'],item['worth_in_usd'],item['in_btc']))
	print ("File {} succesfully saved\n".format(file_name))
	input("Press ENTER to continue")

	
#cell_symbol returns cell adress as string based on row and column - row=1, column=2 = "B1"
def cell_symbol(row,column):
	return "{}{}".format(chr(64+column),row)

def current_value_formula(elements):
	lst=[]
	for i in range(elements):
		lst.append(cell_symbol(i+5,7)+"*"+cell_symbol(i+5,9))
	return ("="+"+".join(lst))


def create_xls(data,money):
	wb=Workbook()
	ws=wb.active

	ws['D4']="Market Cap"
	ws['E4']="Percentage"
	ws['F4']="How much to invest"
	ws['G4']="Unit price (USD)"
	ws['H4']="Units"
	ws['I4']="How much I have now"
	ws['J4']="How much to buy"
	ws['K4']="Balance"

	ws.column_dimensions['C'].width = 15
	ws.column_dimensions['D'].width = 20
	ws.column_dimensions['E'].width = 12
	ws.column_dimensions['F'].width = 20
	ws.column_dimensions['G'].width = 20
	ws.column_dimensions['H'].width = 12
	ws.column_dimensions['I'].width = 20
	ws.column_dimensions['J'].width = 20
	ws.column_dimensions['K'].width = 12

	# filling worsheet with names and values
	#formula_1 calculates how much to buy (target minus how much I have now)
	#formula_2 calculates balance - how many percent actual holdings differ from target
	#formula_3 calculates proportion
	#formula_4 calculates how many units to have
	#formula_5 calculates how much to invest in USD

	for i in range(len(data)):

		formula_1="="+cell_symbol(i+5,8)+"-"+cell_symbol(i+5,9)
		formula_2="="+cell_symbol(i+5,8)+"/"+cell_symbol(i+5,9)+"-1"
		formula_3="="+cell_symbol(i+5,4)+"/D"+str(len(data)+7)
		formula_4="="+cell_symbol(i+5,6)+"/"+cell_symbol(i+5,7)
		formula_5="="+cell_symbol(i+5,5)+"*D"+str(len(data)+9)


		ws.cell(row=(i+5),column=3,value=data[i]['name'])
		ws.cell(row=(i+5),column=4,value=float(data[i]['market_cap_usd'])).number_format='#,##0'
		ws.cell(row=(i+5),column=5,value=formula_3).number_format='0.00%'
		ws.cell(row=(i+5),column=6,value=formula_5).number_format='0.00'
		ws.cell(row=(i+5),column=7,value=float(data[i]['price_usd'])).number_format='0.00'
		ws.cell(row=(i+5),column=8,value=formula_4).number_format='0.000'
		ws.cell(row=(i+5),column=9,value=0).number_format='0.000'
		ws.cell(row=(i+5),column=10,value=formula_1).number_format='0.000'
		ws.cell(row=(i+5),column=11,value=formula_2).number_format='0.00%'
		

	#formula_9 calculates total market cap

	formula_9='=SUM(D5:D'+str(len(data)+4)
	ws.cell(row=len(data)+7,column=3,value='Total:')
	ws.cell(row=len(data)+7,column=4,value=formula_9).number_format='#,##0'
	ws.cell(row=len(data)+9,column=3,value='Money:')
	ws.cell(row=len(data)+9,column=4,value=money).number_format='0.00'
	
	# formula_6 calculates current value

	formula_6=current_value_formula(len(data))
	ws.cell(row=len(data)+11,column=3,value='Current value:')
	ws.cell(row=len(data)+11,column=4,value=formula_6).number_format='0.00'

	
	while True:
		file_name=input('Save portfolio as...(program will add Excel extension .xlsx) : ')
		file_name+=".xlsx"
		try:
			wb.save(file_name)
		except:
			print("File name error!")
		else:
			break
	
	print ("File {} succesfully saved\n".format(file_name))
	input("Press ENTER to continue")
	
	
def load_json():
	url="https://api.coinmarketcap.com/v1/ticker/"
	response = urllib.request.urlopen(url).read()
	data = json.loads(response.decode('utf-8'))
	return data

def find_rank(name): 
	for item in data:
		if item['name']==name:
			return int(item['rank'])
	raise ("Wrong cryptocurrency name in the spreadsheet: {}".format(name))

def update_spreadsheet():
	while True:
		file_name=input("Name of the file to be updated? ")
		try:
			wb=load_workbook(file_name)
		except:
			print("Error! Check name or the type of the file and try again")
		else:
			break
	ws=wb.active
	row_count=ws.max_row

	# Knowing row_count we can calculate how many cryptocurrencies are included in the spreadsheet
	# There are 11 rows not used for storing cryptocurrency data

	crypto_count=row_count-11
	
	#new_list is a sorted list of ranks of cryptocurrencies

	new_list=[]
	for i in range(crypto_count):
		coin_name=ws.cell(row=i+5,column=3).value
		new_list.append(find_rank(coin_name))
		new_list=sorted(new_list)
	
	new_data=[]
	for i in new_list:
		new_data.append(data[i-1])
	
	for i in range(crypto_count):
		ws.cell(row=(i+5),column=4,value=float(new_data[i]['market_cap_usd']))
		ws.cell(row=(i+5),column=7,value=float(new_data[i]['price_usd']))

	print("Your spreadhseet has been updated with current market cap and prices")
	wb.save(file_name)
	print("File {} saved!".format(file_name))	

	input("Press ENTER")
	
# starting main program

os.system('cls')	
print ("\n\n  Reading data from coinmarketcap.com...")
data=load_json()	
while True:
	os.system('cls')
	display_menu()
	
	option=input("\nChoose option: ")
	if option=='3': break
	elif option=='1': choose_portfolio()
	elif option=='2': update_spreadsheet()
	else:
		print ("No such option!")
		input ("Press ENTER to continue")

print("\nThank you for using my program! To the moon! ;)\n")


